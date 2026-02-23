"""Document scanner: extracts text from PDF, images, and Excel."""
from __future__ import annotations

import io
from typing import Any
from pathlib import Path

import pytesseract
from pdf2image import convert_from_path
from PIL import Image
from pypdf import PdfReader

# Excel support
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

ALLOWED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".webp", ".xlsx", ".xls"}


class DocumentScannerAgent:
    def scan_to_markdown(self, path: Path) -> str:
        """Extract text from a file. Supports PDF, images, and Excel."""
        suf = path.suffix.lower()
        if suf == ".pdf":
            return self._scan_pdf(path)
        if suf in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            return self._scan_image(path)
        if suf in (".xlsx", ".xls"):
            return self._scan_excel(path)
        return ""

    def scan_bytes_to_markdown(self, data: bytes, filename: str) -> str:
        """Extract text from bytes (e.g. uploaded file)."""
        suf = Path(filename).suffix.lower()
        if suf == ".pdf":
            return self._scan_pdf_bytes(data)
        if suf in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            return self._scan_image_bytes(data)
        if suf in (".xlsx", ".xls"):
            return self._scan_excel_bytes(data)
        return ""

    def _scan_pdf(self, pdf_path: Path) -> str:
        reader = PdfReader(str(pdf_path))
        chunks: list[str] = []
        for idx, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if text:
                chunks.append(f"# Page {idx}\n\n{text}")

        if chunks:
            return "\n\n".join(chunks)

        images = convert_from_path(str(pdf_path), dpi=200)
        ocr_chunks: list[str] = []
        for idx, image in enumerate(images, start=1):
            text = pytesseract.image_to_string(image).strip()
            ocr_chunks.append(f"# Page {idx}\n\n{text}")
        return "\n\n".join(ocr_chunks)

    def _scan_pdf_bytes(self, data: bytes) -> str:
        reader = PdfReader(io.BytesIO(data))
        chunks: list[str] = []
        for idx, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if text:
                chunks.append(f"# Page {idx}\n\n{text}")

        if chunks:
            return "\n\n".join(chunks)

        # Fallback: write to temp and use OCR (pdf2image needs a file path)
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(data)
            temp_path = Path(f.name)
        try:
            images = convert_from_path(str(temp_path), dpi=200)
            ocr_chunks = []
            for idx, image in enumerate(images, start=1):
                text = pytesseract.image_to_string(image).strip()
                ocr_chunks.append(f"# Page {idx}\n\n{text}")
            return "\n\n".join(ocr_chunks)
        finally:
            temp_path.unlink(missing_ok=True)

    def _scan_image(self, image_path: Path) -> str:
        text = pytesseract.image_to_string(str(image_path)).strip()
        return f"# Page 1\n\n{text}" if text else ""

    def _scan_image_bytes(self, data: bytes) -> str:
        image = Image.open(io.BytesIO(data))
        text = pytesseract.image_to_string(image).strip()
        return f"# Page 1\n\n{text}" if text else ""

    def _scan_excel(self, excel_path: Path) -> str:
        if not OPENPYXL_AVAILABLE:
            return "# Excel\n\n(openpyxl not installed)"
        return self._excel_to_markdown(openpyxl.load_workbook(excel_path, read_only=True, data_only=True))

    def _scan_excel_bytes(self, data: bytes) -> str:
        if not OPENPYXL_AVAILABLE:
            return "# Excel\n\n(openpyxl not installed)"
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return self._excel_to_markdown(wb)

    def _excel_to_markdown(self, wb: Any) -> str:
        chunks: list[str] = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                chunks.append(f"# Sheet: {sheet_name}\n\n" + "\n".join(rows))
        return "\n\n".join(chunks) if chunks else ""
